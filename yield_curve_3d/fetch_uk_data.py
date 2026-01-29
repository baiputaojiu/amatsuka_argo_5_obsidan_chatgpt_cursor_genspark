"""
英国国債（Gilt）イールドカーブを BoE / DMO から取得し、
uk_yield_curve.csv に保存する。

利用元:
- Bank of England: https://www.bankofengland.co.uk/statistics/yield-curves
  - 自動: 当月分のみ ZIP 内 xlsx
- DMO D4H（1998年4月〜現在、月次・5Y/10Y/30Y/50Y）:
  https://www.dmo.gov.uk/data/ExportReport?reportCode=D4H
  - サイトは CAPTCHA のため自動取得不可。手動で Excel を DL し、
    python fetch_uk_data.py ダウンロードした.xlsx で正規化。

使い方:
  python fetch_uk_data.py
      # BoE から当月分を自動取得
  python fetch_uk_data.py ダウンロードした.csv または .xlsx
      # 手動DL（BoE アーカイブ or DMO D4H）を正規化して保存
"""

import sys
from io import BytesIO, StringIO
from pathlib import Path
import zipfile

import pandas as pd

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
OUTPUT_CSV = DATA_DIR / "uk_yield_curve.csv"

# BoE 最新イールドカーブ（ZIP 内に xlsx; 名目国債 GLC Nominal を優先）
BOE_ZIP_URL = "https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/latest-yield-curve-data.zip"
# 名目国債の xlsx を優先して使用
BOE_NOMINAL_XLSX = "GLC Nominal daily data current month.xlsx"

# app.py が期待する形式: Date + 残存期間列（例: 5Y, 10Y, 30Y）


def _download_bytes(url: str, timeout: int = 120) -> bytes:
    try:
        import urllib.request
        import ssl
        req = urllib.request.Request(
            url,
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "Accept": "*/*",
            },
        )
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        with urllib.request.urlopen(req, timeout=timeout, context=ctx) as res:
            return res.read()
    except Exception as e:
        raise RuntimeError(f"ダウンロードに失敗しました: {url}") from e


def _normalize_uk_csv(df: pd.DataFrame, since_year: int = 2000) -> pd.DataFrame:
    """日付列を Date に、残存期間列を数値化。since_year 以降に絞る。"""
    df = df.copy()
    date_col = df.columns[0]
    df = df.rename(columns={date_col: "Date"})
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df = df.dropna(subset=["Date"])
    df = df[df["Date"].dt.year >= since_year]
    maturity_cols = [c for c in df.columns if c != "Date"]
    for c in maturity_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    df = df.sort_values("Date", ascending=False)
    df["Date"] = df["Date"].dt.strftime("%Y-%m-%d")
    return df


# DMO D4H: 月次・5Y/10Y/30Y/50Y。列名は "5 year", "10 year" 等の可能性
D4H_MATURITY_MAP = {"5": "5Y", "10": "10Y", "20": "20Y", "30": "30Y", "50": "50Y"}


def _detect_and_normalize_d4h(df: pd.DataFrame) -> pd.DataFrame | None:
    """
    DMO D4H 形式の DataFrame かどうか判定し、Date + 5Y, 10Y, 30Y, 50Y に正規化する。
    D4H は 1998年4月〜、月次で 5Y/10Y/30Y（2005年6月〜50Y）。列名は "5 year" 等。
    """
    if df.empty or len(df.columns) < 3:
        return None
    df = df.copy()
    # 先頭列を日付に
    first = df.columns[0]
    df = df.rename(columns={first: "Date"})
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df = df.dropna(subset=["Date"])
    if len(df) == 0:
        return None
    # 残り列を残存期間にマッピング（5 year -> 5Y 等）
    new_cols = {"Date": df["Date"]}
    for c in df.columns:
        if c == "Date":
            continue
        cstr = str(c).strip().lower()
        mapped = None
        for k, v in D4H_MATURITY_MAP.items():
            if k in cstr or cstr.startswith(k + " ") or cstr == k:
                mapped = v
                break
        if mapped is None and cstr in ("5y", "10y", "20y", "30y", "50y"):
            mapped = cstr.upper()
        if mapped:
            new_cols[mapped] = pd.to_numeric(df[c], errors="coerce")
    if len(new_cols) < 3:
        return None
    order = ["Date"] + [c for c in ("5Y", "10Y", "20Y", "30Y", "50Y") if c in new_cols]
    out = pd.DataFrame({k: new_cols[k] for k in order})
    out = out.dropna(subset=["Date"])
    out = out[out["Date"].dt.year >= 1998]
    out = out.sort_values("Date", ascending=False)
    out["Date"] = out["Date"].dt.strftime("%Y-%m-%d")
    return out


def _read_df_from_zip(z: zipfile.ZipFile, name: str, since_year: int) -> pd.DataFrame | None:
    """ZIP 内の CSV または xlsx を読み、正規化して返す。"""
    try:
        raw = z.read(name)
        if name.lower().endswith(".csv"):
            text = raw.decode("utf-8", errors="replace")
            df = pd.read_csv(StringIO(text))
        elif name.lower().endswith(".xlsx"):
            # BoE xlsx: 結合セルで pandas が1列に読むため、openpyxl でセル単位に読み DataFrame を構築
            import openpyxl
            wb = openpyxl.load_workbook(BytesIO(bytes(raw)), read_only=False, data_only=True)
            ws = wb.active
            max_row, max_col = ws.max_row, ws.max_column
            if max_row < 7 or max_col < 2:
                wb.close()
                return None
            # 4行目をヘッダー、6行目以降をデータ
            headers = [ws.cell(row=4, column=c).value for c in range(1, max_col + 1)]
            if headers[0] is None or not isinstance(headers[0], (int, float)):
                headers[0] = "Date"
            rows = []
            for r in range(6, max_row + 1):
                row = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
                if row[0] is None:
                    continue
                rows.append(row)
            wb.close()
            df = pd.DataFrame(rows, columns=headers)
        else:
            return None
        if df.empty or len(df.columns) < 2:
            return None
        df = _normalize_uk_csv(df, since_year=since_year)
        return df if len(df) > 0 else None
    except Exception:
        return None


def _fetch_from_boe_zip(since_year: int = 2000) -> pd.DataFrame | None:
    """BoE の ZIP を取得し、名目国債 xlsx または CSV を正規化して返す。"""
    try:
        raw = _download_bytes(BOE_ZIP_URL)
    except Exception:
        return None
    try:
        z = zipfile.ZipFile(BytesIO(raw), "r")
    except Exception:
        return None
    names = z.namelist()
    # 名目国債 xlsx を優先
    nominal_name = next((n for n in names if "Nominal" in n and n.endswith(".xlsx")), None)
    if nominal_name:
        df = _read_df_from_zip(z, nominal_name, since_year)
        if df is not None:
            return df
    for name in names:
        if name.lower().endswith(".csv"):
            df = _read_df_from_zip(z, name, since_year)
            if df is not None:
                return df
        if name.lower().endswith(".xlsx"):
            df = _read_df_from_zip(z, name, since_year)
            if df is not None:
                return df
    return None


def main() -> None:
    since_year = 2000
    if len(sys.argv) >= 2:
        path = Path(sys.argv[1]).resolve()
        if not path.exists():
            print(f"ファイルが見つかりません: {path}")
            raise SystemExit(1)
        print(f"読み込み中: {path}")
        if path.suffix.lower() == ".xlsx":
            df_raw = pd.read_excel(path, engine="openpyxl", sheet_name=0)
            # DMO D4H 形式（1998年4月〜月次・5Y/10Y/30Y/50Y）ならそのまま正規化
            df_d4h = _detect_and_normalize_d4h(df_raw)
            if df_d4h is not None and len(df_d4h) > 0:
                df = df_d4h
                print("  → DMO D4H 形式として処理しました（1998年4月〜）。")
            else:
                df = _normalize_uk_csv(df_raw, since_year=since_year)
        else:
            df = pd.read_csv(path)
            df = _normalize_uk_csv(df, since_year=since_year)
    else:
        print("英国イールドカーブ（2000年〜）を取得しています...")
        df = _fetch_from_boe_zip(since_year=since_year)
        if df is None or len(df) == 0:
            print(
                "自動ダウンロードに失敗しました。\n"
                "手動で以下からデータを取得し、\n"
                "  python fetch_uk_data.py ダウンロードしたファイル.csv または .xlsx\n"
                "で正規化して保存できます。\n"
                f"  BoE: {BOE_ZIP_URL}\n"
                "  DMO: https://www.dmo.gov.uk/data/ExportReport?reportCode=D4H"
            )
            raise SystemExit(1)

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    df.to_csv(OUTPUT_CSV, index=False)
    n = len(df)
    if n > 0:
        print(f"保存しました: {OUTPUT_CSV}（{n} 行、{df['Date'].iloc[-1]} 〜 {df['Date'].iloc[0]}）")
    else:
        print("データが 0 行でした。")


if __name__ == "__main__":
    main()
